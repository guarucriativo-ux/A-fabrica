"""
Adiciona uma linha na aba 'Lançamentos' do Controle_Financeiro_Geral.xlsx,
preservando formatação. Uso interno do Claude — não precisa rodar manualmente.

Uso:
python3 add_lancamento.py "AAAA-MM-DD" "Conta" "Tipo" "Categoria" "Descrição" Valor "Forma de Pagamento"
"""
import sys, openpyxl
from openpyxl.styles import Font, Border, Side
from datetime import datetime

PATH = "Controle_Financeiro_Geral.xlsx"
FONT = "Arial"
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def add(data_str, conta, tipo, categoria, descricao, valor, forma):
    wb = openpyxl.load_workbook(PATH)
    ws = wb["Lançamentos"]
    # acha a primeira linha vazia (coluna A) depois dos exemplos
    r = 2
    while ws.cell(row=r, column=1).value not in (None, ""):
        r += 1
    data = datetime.strptime(data_str, "%Y-%m-%d").date()
    vals = [data, conta, tipo, categoria, descricao, float(valor), forma]
    for col, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=col, value=v)
        c.font = Font(name=FONT)
        c.border = BORDER
        if col == 1:
            c.number_format = "DD/MM/AAAA"
        if col == 6:
            c.number_format = "R$ #,##0.00"
    ws.cell(row=r, column=8, value=f"=IF(A{r}=\"\",\"\",DATE(YEAR(A{r}),MONTH(A{r}),1))")
    ws.cell(row=r, column=8).number_format = "MMM/AAAA"
    ws.cell(row=r, column=8).font = Font(name=FONT)
    ws.cell(row=r, column=8).border = BORDER
    wb.save(PATH)
    print(f"Lançado na linha {r}: {data_str} | {conta} | {tipo} | {categoria} | {descricao} | R${valor} | {forma}")

if __name__ == "__main__":
    add(*sys.argv[1:8])
