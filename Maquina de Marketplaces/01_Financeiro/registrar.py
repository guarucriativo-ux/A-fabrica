"""
Registra qualquer evento de negócio no ERP_Guaru_Estudio.xlsx num comando só:
venda, fornecedor novo, cliente novo, conta a pagar, conta a receber.
Sempre recalcula e copia pro destino final automaticamente — mesma lógica do
lancar.py (que cobre só Lançamentos Pessoais).

USO:
  python3 registrar.py venda "<cliente>" "<produto>" <qtd> <valor_unit> "<canal>" [--status Confirmado] [--nf ""] [--data AAAA-MM-DD]
  python3 registrar.py fornecedor "<nome>" "<o que fornece>" "<contato>" "<condições>" "<usado em>" [--status Ativo]
  python3 registrar.py cliente "<nome>" "<contato>" "<canal de origem>" [--obs ""] [--data AAAA-MM-DD]
  python3 registrar.py contapagar "<descrição>" "<categoria>" <valor> "<vencimento>" [--status Pendente] [--data AAAA-MM-DD]
  python3 registrar.py contareceber "<cliente/canal>" "<pedido vinculado>" <valor> [--status Pendente] [--data AAAA-MM-DD]

Exemplos:
  python3 registrar.py venda "Maria Silva" "Kit Doceria" 2 34.90 Shopee
  python3 registrar.py fornecedor "Scan Revenda" "Adesivo, vinil, DTF, acrílico" "Balcão Santos" "Retirada presencial" "Adesivos diversos"
  python3 registrar.py cliente "João Pedro" "(13) 99999-0000" WhatsApp
  python3 registrar.py contapagar "Lote piloto GIV" "Compra de Mercadoria/Insumo" 54.99 Único
  python3 registrar.py contareceber "Pedido #102 - Shopee" "Kit Floricultura" 59.90
"""
import sys
import argparse
import subprocess
import shutil
import os
from datetime import date
import openpyxl

ERP_PATH = "ERP_Guaru_Estudio.xlsx"
RECALC_SCRIPT = "/sessions/friendly-keen-albattani/mnt/.claude/skills/xlsx/scripts/recalc.py"
DEST_PATH = "/sessions/friendly-keen-albattani/mnt/guaru_2.0_itsnow/Maquina de Marketplaces/01_Financeiro/ERP_Guaru_Estudio.xlsx"


def next_empty_row(ws, col=1, start=2):
    r = start
    while ws.cell(row=r, column=col).value is not None:
        r += 1
    return r


def finish(wb, msg):
    wb.save(ERP_PATH)
    print(msg)
    result = subprocess.run(["python3", RECALC_SCRIPT, ERP_PATH, "90"], capture_output=True, text=True)
    print(result.stdout)
    if "errors_found" in result.stdout:
        print("ATENCAO: recalc encontrou erros, revisar antes de copiar.")
        sys.exit(1)
    if os.path.abspath(ERP_PATH) != os.path.abspath(DEST_PATH):
        shutil.copy(ERP_PATH, DEST_PATH)
        print(f"Copiado para {DEST_PATH}")


def cmd_venda(a):
    wb = openpyxl.load_workbook(ERP_PATH)
    ws = wb["Vendas"]
    r = next_empty_row(ws)
    d = date.fromisoformat(a.data) if a.data else date.today()
    ws.cell(row=r, column=1, value=d).number_format = "dd/mm/yyyy"
    ws.cell(row=r, column=2, value=a.cliente)
    ws.cell(row=r, column=3, value=a.produto)
    ws.cell(row=r, column=4, value=a.qtd)
    ws.cell(row=r, column=5, value=a.valor_unit).number_format = "R$ #,##0.00"
    ws.cell(row=r, column=6, value=f'=IF(OR(D{r}="",E{r}=""),"",D{r}*E{r})').number_format = "R$ #,##0.00"
    ws.cell(row=r, column=7, value=a.canal)
    ws.cell(row=r, column=8, value=a.status)
    ws.cell(row=r, column=9, value=a.nf or "")
    finish(wb, f"Venda lançada na linha {r}: {d} | {a.cliente} | {a.produto} x{a.qtd} | R${a.valor_unit:.2f}/un | {a.canal} | {a.status}")


def cmd_fornecedor(a):
    wb = openpyxl.load_workbook(ERP_PATH)
    ws = wb["Fornecedores"]
    r = next_empty_row(ws)
    ws.cell(row=r, column=1, value=a.nome)
    ws.cell(row=r, column=2, value=a.fornece)
    ws.cell(row=r, column=3, value=a.contato)
    ws.cell(row=r, column=4, value=a.condicoes)
    ws.cell(row=r, column=5, value=a.usado_em)
    ws.cell(row=r, column=6, value=a.status)
    finish(wb, f"Fornecedor lançado na linha {r}: {a.nome} | {a.fornece} | {a.status}")


def cmd_cliente(a):
    wb = openpyxl.load_workbook(ERP_PATH)
    ws = wb["Clientes"]
    r = next_empty_row(ws)
    d = date.fromisoformat(a.data) if a.data else date.today()
    ws.cell(row=r, column=1, value=a.nome)
    ws.cell(row=r, column=2, value=a.contato)
    ws.cell(row=r, column=3, value=a.canal)
    ws.cell(row=r, column=4, value=d).number_format = "dd/mm/yyyy"
    ws.cell(row=r, column=5, value=f"=SUMIFS(Vendas!F:F,Vendas!B:B,A{r})").number_format = "R$ #,##0.00"
    ws.cell(row=r, column=6, value=a.obs or "")
    finish(wb, f"Cliente lançado na linha {r}: {a.nome} | {a.contato} | {a.canal}")


def cmd_contapagar(a):
    wb = openpyxl.load_workbook(ERP_PATH)
    ws = wb["Contas a Pagar"]
    r = next_empty_row(ws)
    d = date.fromisoformat(a.data) if a.data else date.today()
    ws.cell(row=r, column=1, value=d).number_format = "dd/mm/yyyy"
    ws.cell(row=r, column=2, value=a.descricao)
    ws.cell(row=r, column=3, value=a.categoria)
    ws.cell(row=r, column=4, value=a.valor).number_format = "R$ #,##0.00"
    ws.cell(row=r, column=5, value=a.vencimento)
    ws.cell(row=r, column=6, value=a.status)
    finish(wb, f"Conta a Pagar lançada na linha {r}: {a.descricao} | R${a.valor:.2f} | vence {a.vencimento} | {a.status}")


def cmd_contareceber(a):
    wb = openpyxl.load_workbook(ERP_PATH)
    ws = wb["Contas a Receber"]
    r = next_empty_row(ws)
    d = date.fromisoformat(a.data) if a.data else date.today()
    ws.cell(row=r, column=1, value=d).number_format = "dd/mm/yyyy"
    ws.cell(row=r, column=2, value=a.cliente_canal)
    ws.cell(row=r, column=3, value=a.pedido)
    ws.cell(row=r, column=4, value=a.valor).number_format = "R$ #,##0.00"
    ws.cell(row=r, column=5, value=a.status)
    finish(wb, f"Conta a Receber lançada na linha {r}: {a.cliente_canal} | {a.pedido} | R${a.valor:.2f} | {a.status}")


def main():
    p = argparse.ArgumentParser()
    sub = p.add_subparsers(dest="cmd", required=True)

    pv = sub.add_parser("venda")
    pv.add_argument("cliente")
    pv.add_argument("produto")
    pv.add_argument("qtd", type=int)
    pv.add_argument("valor_unit", type=float)
    pv.add_argument("canal")
    pv.add_argument("--status", default="Confirmado")
    pv.add_argument("--nf", default="")
    pv.add_argument("--data", default=None)
    pv.set_defaults(func=cmd_venda)

    pf = sub.add_parser("fornecedor")
    pf.add_argument("nome")
    pf.add_argument("fornece")
    pf.add_argument("contato")
    pf.add_argument("condicoes")
    pf.add_argument("usado_em")
    pf.add_argument("--status", default="Ativo")
    pf.set_defaults(func=cmd_fornecedor)

    pc = sub.add_parser("cliente")
    pc.add_argument("nome")
    pc.add_argument("contato")
    pc.add_argument("canal")
    pc.add_argument("--obs", default="")
    pc.add_argument("--data", default=None)
    pc.set_defaults(func=cmd_cliente)

    pp = sub.add_parser("contapagar")
    pp.add_argument("descricao")
    pp.add_argument("categoria")
    pp.add_argument("valor", type=float)
    pp.add_argument("vencimento")
    pp.add_argument("--status", default="Pendente")
    pp.add_argument("--data", default=None)
    pp.set_defaults(func=cmd_contapagar)

    pr = sub.add_parser("contareceber")
    pr.add_argument("cliente_canal")
    pr.add_argument("pedido")
    pr.add_argument("valor", type=float)
    pr.add_argument("--status", default="Pendente")
    pr.add_argument("--data", default=None)
    pr.set_defaults(func=cmd_contareceber)

    args = p.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
