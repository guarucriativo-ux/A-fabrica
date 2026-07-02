"""
Lança uma transação em Lançamentos Pessoal do ERP_Guaru_Estudio.xlsx,
atualiza o Saldo Atual no Dashboard automaticamente, recalcula e copia pra
pasta final — tudo em UM comando, sem precisar escrever script novo cada vez.

Uso (linha de comando, a partir desta pasta):
    python3 lancar.py <Entrada|Saida> <valor> <categoria> "<descrição>" [--conta Pessoal] [--forma "Pix"] [--data AAAA-MM-DD]

Exemplos:
    python3 lancar.py Saida 40 Transporte "Gasolina"
    python3 lancar.py Entrada 100 "Recebível de Venda Anterior" "Quitanda - Pix diário"
    python3 lancar.py Saida 32.99 "Compra de Mercadoria/Insumo" "GIV Online" --conta "Máquina de Marketplaces"

Por padrão: conta=Pessoal, forma=Não informado, data=hoje.
Só afeta o Saldo Atual do Dashboard quando conta=Pessoal (negócio não usa esse campo).
"""
import sys
import argparse
import subprocess
from datetime import date
import openpyxl

ERP_PATH = "ERP_Guaru_Estudio.xlsx"
RECALC_SCRIPT = "/sessions/friendly-keen-albattani/mnt/.claude/skills/xlsx/scripts/recalc.py"
DEST_PATH = "/sessions/friendly-keen-albattani/mnt/guaru_2.0_itsnow/Maquina de Marketplaces/01_Financeiro/ERP_Guaru_Estudio.xlsx"


def main():
    p = argparse.ArgumentParser()
    p.add_argument("tipo", choices=["Entrada", "Saida", "Saída"])
    p.add_argument("valor", type=float)
    p.add_argument("categoria")
    p.add_argument("descricao")
    p.add_argument("--conta", default="Pessoal")
    p.add_argument("--forma", default="Não informado")
    p.add_argument("--data", default=None, help="AAAA-MM-DD, padrão hoje")
    args = p.parse_args()

    tipo = "Saída" if args.tipo in ("Saida", "Saída") else "Entrada"
    data_lanc = date.fromisoformat(args.data) if args.data else date.today()

    wb = openpyxl.load_workbook(ERP_PATH)
    ws = wb["Lançamentos Pessoal"]
    r = 2
    while ws.cell(row=r, column=1).value is not None:
        r += 1

    ws.cell(row=r, column=1, value=data_lanc).number_format = "dd/mm/yyyy"
    ws.cell(row=r, column=2, value=args.conta)
    ws.cell(row=r, column=3, value=tipo)
    ws.cell(row=r, column=4, value=args.categoria)
    ws.cell(row=r, column=5, value=args.descricao)
    ws.cell(row=r, column=6, value=args.valor)
    ws.cell(row=r, column=7, value=args.forma)
    ws.cell(row=r, column=8, value=f'=IF(A{r}="","",DATE(YEAR(A{r}),MONTH(A{r}),1))')

    if args.conta == "Pessoal":
        ws_d = wb["Dashboard"]
        saldo_atual = ws_d["B37"].value or 0
        novo_saldo = saldo_atual + args.valor if tipo == "Entrada" else saldo_atual - args.valor
        ws_d["B37"] = novo_saldo
        ws_d["C37"] = f"Após {tipo.lower()} de R${args.valor:.2f} ({args.descricao}) em {data_lanc.strftime('%d/%m')}."
        print(f"Saldo Atual: R${saldo_atual:.2f} -> R${novo_saldo:.2f}")

    wb.save(ERP_PATH)
    print(f"Lançado na linha {r}: {data_lanc} | {args.conta} | {tipo} | {args.categoria} | {args.descricao} | R${args.valor:.2f} | {args.forma}")

    # recalcula
    result = subprocess.run(["python3", RECALC_SCRIPT, ERP_PATH, "90"], capture_output=True, text=True)
    print(result.stdout)
    if "errors_found" in result.stdout:
        print("ATENCAO: recalc encontrou erros, revisar antes de copiar.")
        sys.exit(1)

    # copia pra destino (se já não for o mesmo arquivo)
    import shutil, os
    if os.path.abspath(ERP_PATH) != os.path.abspath(DEST_PATH):
        shutil.copy(ERP_PATH, DEST_PATH)
        print(f"Copiado para {DEST_PATH}")
    else:
        print("Arquivo já está no destino final.")


if __name__ == "__main__":
    main()
