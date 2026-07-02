import re, json, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"; PURPLE = "5B3FA0"; LIGHT = "F2F0FA"; WHITE = "FFFFFF"
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

SRC_JS = sys.argv[1] if len(sys.argv) > 1 else "/sessions/friendly-keen-albattani/mnt/guaru_2.0_itsnow/Tabela de preços/js/dados_produtos.js"
ERP_PATH = sys.argv[2] if len(sys.argv) > 2 else "/sessions/friendly-keen-albattani/mnt/outputs/erp/ERP_Guaru_Estudio.xlsx"

with open(SRC_JS, "r", encoding="utf-8") as f:
    content = f.read()

pattern = re.compile(r'\{name:"([^"]+)",tier:"([^"]+)",gram:"([^"]+)",qty:(\d+),custo:(\d+(?:\.\d+)?),venda:(\d+(?:\.\d+)?),corte:(true|false)\}')
rows = []
for m in pattern.finditer(content):
    rows.append({
        "name": m.group(1), "tier": m.group(2), "gram": m.group(3),
        "qty": int(m.group(4)), "custo": float(m.group(5)), "venda": float(m.group(6)),
    })

by_name = {}
for r in rows:
    d = by_name.setdefault(r["name"], {"min_custo": float("inf"), "max_custo": 0, "min_venda": float("inf"), "max_venda": 0, "count": 0, "tiers": set()})
    d["min_custo"] = min(d["min_custo"], r["custo"])
    d["max_custo"] = max(d["max_custo"], r["custo"])
    d["min_venda"] = min(d["min_venda"], r["venda"])
    d["max_venda"] = max(d["max_venda"], r["venda"])
    d["count"] += 1
    d["tiers"].add(r["tier"])

categoria_map = {
    "Couchê": "Guaru Estúdio — Cartão", "Verniz": "Guaru Estúdio — Cartão", "Laminação": "Guaru Estúdio — Cartão",
    "Hot Stamping": "Guaru Estúdio — Cartão", "Adesivo": "Guaru Estúdio — Adesivo", "Sticker": "Guaru Estúdio — Adesivo",
    "Rótulo": "Guaru Estúdio — Adesivo", "Lacre": "Guaru Estúdio — Adesivo",
}
def categoria_de(nome):
    for k, v in categoria_map.items():
        if k.lower() in nome.lower():
            return v
    return "Guaru Estúdio — Outro"

wb = openpyxl.load_workbook(ERP_PATH)
ws = wb["Produtos"]
for row in ws.iter_rows(min_row=1, max_row=60):
    for cell in row:
        cell.value = None

def header_row(ws, row, headers, widths=None):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=FONT, bold=True, color=WHITE)
        c.fill = PatternFill("solid", start_color=PURPLE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        if widths: ws.column_dimensions[get_column_letter(i)].width = widths[i-1]

header_row(ws, 1, ["Produto", "Categoria", "Faixa Custo", "Faixa Venda", "Margem %", "Nº Variações", "Fornecedor", "Status", "Observação"],
           [32, 22, 16, 16, 11, 13, 18, 12, 30])

r = 2
for name, d in sorted(by_name.items()):
    ws.cell(row=r, column=1, value=name).font = Font(name=FONT)
    ws.cell(row=r, column=2, value=categoria_de(name)).font = Font(name=FONT)
    ws.cell(row=r, column=3, value=f"R$ {d['min_custo']:.2f} – R$ {d['max_custo']:.2f}").font = Font(name=FONT)
    ws.cell(row=r, column=4, value=f"R$ {d['min_venda']:.2f} – R$ {d['max_venda']:.2f}").font = Font(name=FONT)
    ws.cell(row=r, column=5, value="50.0%").font = Font(name=FONT)
    ws.cell(row=r, column=6, value=d["count"]).font = Font(name=FONT)
    ws.cell(row=r, column=7, value="GIV Online + CMB").font = Font(name=FONT)
    ws.cell(row=r, column=8, value="Ativo").font = Font(name=FONT)
    ws.cell(row=r, column=9, value="Sincronizado de dados_produtos.js (site)").font = Font(name=FONT, italic=True, size=9, color="808080")
    for col in range(1, 10):
        ws.cell(row=r, column=col).border = BORDER
        if r % 2 == 1:
            ws.cell(row=r, column=col).fill = PatternFill("solid", start_color=LIGHT)
    r += 1

r += 1
ws.cell(row=r, column=1, value="— Linha Máquina de Marketplaces (kits fracionados) —").font = Font(name=FONT, bold=True, italic=True, color=PURPLE)
for col in range(1, 10): ws.cell(row=r, column=col).border = BORDER
r += 1

mm_kits = [
    ["Kit Acabamento Doceria", "Máquina de Marketplaces", "R$ 24,35", "R$ 34,90", "30.2%", "1", "GIV Online + CMB", "Ativo", "Já listado no site"],
    ["Kit Etiqueta Bijuteria", "Máquina de Marketplaces", "R$ 30,32", "R$ 39,90", "24.0%", "1", "GIV Online + CMB", "Ativo", "Copy pronta"],
    ["Kit Delivery", "Máquina de Marketplaces", "R$ 32,18", "R$ 42,90", "25.0%", "1", "GIV Online + CMB", "Ativo", "Copy pronta"],
    ["Kit Floricultura", "Máquina de Marketplaces", "R$ 49,71", "R$ 59,90", "17.0%", "1", "GIV Online + CMB", "Ativo", "Copy pronta"],
    ["Avental Personalizado DTF", "Guaru Estúdio — Têxtil", "R$ 13,00", "R$ 34,90", "62.7%", "1", "Scan Revenda + Silkado", "Em avaliação", "Ver Radar de Oportunidades"],
]
for row in mm_kits:
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = Font(name=FONT, italic=(i == 9), size=9 if i == 9 else 11, color="808080" if i == 9 else "000000")
        c.border = BORDER
        if r % 2 == 1:
            c.fill = PatternFill("solid", start_color=LIGHT)
    r += 1

last_data_row = r - 1
for rr in range(r, r + 10):
    for col in range(1, 10):
        ws.cell(row=rr, column=col).border = BORDER
        ws.cell(row=rr, column=col).font = Font(name=FONT)

dv_cat = DataValidation(type="list", formula1='"Guaru Estúdio — Cartão,Guaru Estúdio — Adesivo,Guaru Estúdio — Têxtil,Guaru Estúdio — Outro,Máquina de Marketplaces"', allow_blank=True)
ws.add_data_validation(dv_cat); dv_cat.add(f"B2:B{r+10}")
dv_status = DataValidation(type="list", formula1='"Ativo,Em avaliação,Descontinuado"', allow_blank=True)
ws.add_data_validation(dv_status); dv_status.add(f"H2:H{r+10}")

wb.save(ERP_PATH)
print(json.dumps({"produtos_site": len(by_name), "kits_mm": len(mm_kits), "last_row": last_data_row}))
